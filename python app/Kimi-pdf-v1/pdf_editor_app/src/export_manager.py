"""
Export Manager Module - Handles exporting PDFs to various formats
"""

import os
from pathlib import Path
from typing import Optional, Callable, List, Dict
import io


class ExportManager:
    """Manages exporting PDFs to different formats."""
    
    def __init__(self):
        pass
        
    def export_to_word(self, pdf_engine, output_path: str, 
                       progress_callback: Optional[Callable[[int], None]] = None):
        """Export PDF to Word document."""
        try:
            from docx import Document
            from docx.shared import Inches, Pt
        except ImportError:
            raise RuntimeError("python-docx not installed. Install with: pip install python-docx")
            
        doc = Document()
        
        for i in range(pdf_engine.page_count):
            page = pdf_engine.get_page(i)
            if page:
                # Extract text
                text = page.get_text()
                
                # Add to document
                if i > 0:
                    doc.add_page_break()
                    
                # Split by paragraphs
                paragraphs = text.split('\n\n')
                for para_text in paragraphs:
                    if para_text.strip():
                        doc.add_paragraph(para_text.strip())
                        
            if progress_callback:
                progress = int((i + 1) / pdf_engine.page_count * 100)
                progress_callback(progress)
                
        doc.save(output_path)
        
    def export_to_excel(self, pdf_engine, output_path: str,
                        progress_callback: Optional[Callable[[int], None]] = None):
        """Export PDF tables to Excel."""
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl")
            
        wb = Workbook()
        ws = wb.active
        ws.title = "PDF Content"
        
        row = 1
        for i in range(pdf_engine.page_count):
            page = pdf_engine.get_page(i)
            if page:
                # Extract text
                text = page.get_text()
                
                # Add page header
                ws.cell(row, 1, f"Page {i + 1}")
                ws.cell(row, 1).font = openpyxl.styles.Font(bold=True)
                row += 1
                
                # Add content
                lines = text.split('\n')
                for line in lines:
                    if line.strip():
                        ws.cell(row, 1, line.strip())
                        row += 1
                        
                row += 1  # Empty row between pages
                
            if progress_callback:
                progress = int((i + 1) / pdf_engine.page_count * 100)
                progress_callback(progress)
                
        wb.save(output_path)
        
    def export_to_images(self, pdf_engine, output_folder: str,
                         format: str = "png",
                         dpi: int = 150,
                         progress_callback: Optional[Callable[[int], None]] = None):
        """Export PDF pages as images."""
        from PIL import Image
        
        os.makedirs(output_folder, exist_ok=True)
        
        base_name = Path(pdf_engine.file_path).stem if pdf_engine.file_path else "page"
        
        for i in range(pdf_engine.page_count):
            page = pdf_engine.get_page(i)
            if page:
                # Render at specified DPI
                zoom = dpi / 72
                pix = page.get_pixmap(zoom=zoom)
                
                # Convert to PIL Image
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                
                # Save
                output_path = os.path.join(output_folder, f"{base_name}_page_{i+1}.{format}")
                img.save(output_path, format.upper())
                
            if progress_callback:
                progress = int((i + 1) / pdf_engine.page_count * 100)
                progress_callback(progress)
                
    def export_to_text(self, pdf_engine, output_path: str,
                       progress_callback: Optional[Callable[[int], None]] = None):
        """Export PDF to plain text."""
        with open(output_path, 'w', encoding='utf-8') as f:
            for i in range(pdf_engine.page_count):
                page = pdf_engine.get_page(i)
                if page:
                    f.write(f"\n{'='*50}\n")
                    f.write(f"Page {i + 1}\n")
                    f.write(f"{'='*50}\n\n")
                    
                    text = page.get_text()
                    f.write(text)
                    f.write('\n')
                    
                if progress_callback:
                    progress = int((i + 1) / pdf_engine.page_count * 100)
                    progress_callback(progress)
                    
    def export_to_html(self, pdf_engine, output_path: str,
                       progress_callback: Optional[Callable[[int], None]] = None):
        """Export PDF to HTML."""
        html_parts = [
            "<!DOCTYPE html>",
            "<html>",
            "<head>",
            "<meta charset='UTF-8'>",
            "<title>Exported PDF</title>",
            "<style>",
            "body { font-family: Arial, sans-serif; margin: 40px; }",
            ".page { border: 1px solid #ccc; padding: 20px; margin-bottom: 20px; }",
            ".page-header { font-weight: bold; margin-bottom: 10px; color: #666; }",
            "</style>",
            "</head>",
            "<body>",
            "<h1>Exported PDF</h1>"
        ]
        
        for i in range(pdf_engine.page_count):
            page = pdf_engine.get_page(i)
            if page:
                html_parts.append(f'<div class="page">')
                html_parts.append(f'<div class="page-header">Page {i + 1}</div>')
                
                text = page.get_text()
                # Escape HTML
                text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                # Convert newlines to <br>
                text = text.replace('\n', '<br>')
                
                html_parts.append(f'<div class="content">{text}</div>')
                html_parts.append('</div>')
                
            if progress_callback:
                progress = int((i + 1) / pdf_engine.page_count * 100)
                progress_callback(progress)
                
        html_parts.extend([
            "</body>",
            "</html>"
        ])
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(html_parts))
            
    def export_page_as_image(self, pdf_engine, page_num: int,
                             output_path: str,
                             format: str = "png",
                             dpi: int = 150):
        """Export a single page as image."""
        from PIL import Image
        
        page = pdf_engine.get_page(page_num)
        if not page:
            raise ValueError(f"Page {page_num} not found")
            
        zoom = dpi / 72
        pix = page.get_pixmap(zoom=zoom)
        
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        img.save(output_path, format.upper())
        
    def export_to_pdf_a(self, pdf_engine, output_path: str,
                        compliance_level: str = "PDF/A-1b"):
        """Export to PDF/A format (archival standard)."""
        # This would use a PDF/A conversion library
        # For now, just save with basic compliance
        pdf_engine.save(output_path)
        
    def export_optimized_for_web(self, pdf_engine, output_path: str):
        """Export PDF optimized for web viewing."""
        # Linearized PDF for fast web viewing
        if pdf_engine.document:
            pdf_engine.document.save(
                output_path,
                garbage=4,
                deflate=True,
                linear=True  # Linearized for web
            )
            
    def export_redacted(self, pdf_engine, output_path: str,
                        redaction_areas: List[Dict]):
        """Export PDF with redacted content."""
        # Apply redactions and export
        for area in redaction_areas:
            page_num = area.get('page', 0)
            rect = area.get('rect')
            
            page = pdf_engine.get_page(page_num)
            if page and rect:
                page.fitz_page.add_redact_annot(rect)
                page.fitz_page.apply_redactions()
                
        pdf_engine.save(output_path)
